import os
import re
import datetime
import logging
import asyncio
from typing import Optional

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters,
)

import config
import db
import llm_ocr

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

# Pre-load EasyOCR reader so fallback is instant
try:
    import easyocr
    import numpy as np
    from PIL import Image
    _EASYOCR_READER = easyocr.Reader(["en"], gpu=False)
    logger.info("EasyOCR pre-loaded")
except Exception as e:
    logger.warning("EasyOCR not available: %s", e)
    _EASYOCR_READER = None

# Conversation states
PHOTO, TYPE, DURATION, CALORIES, DISTANCE, CONFIRM, WAITING_WEIGHT = range(7)

MACHINE_OPTIONS = [
    ["🪜 Stair Master", "🏔️ Incline Treadmill"],
    ["🚲 Indoor Cycling", "💪 Bicep-Tricep Curl"],
    ["🦵 Leg", "🏋️ Strength- Others"],
]

TYPE_MAP = {
    "🪜 Stair Master": "Cardio",
    "🏔️ Incline Treadmill": "Cardio",
    "🚲 Indoor Cycling": "Cardio",
    "💪 Bicep-Tricep Curl": "Strength",
    "🦵 Leg": "Strength",
    "🏋️ Strength- Others": "Strength",
}

# kcal/min estimate for strength machines when user skips calorie entry
# Sources: Reis et al. 2017 (PLoS One), Adeel et al. 2021 (Appl. Sci.),
# StrongerByScience review of João et al. 2022.
STRENGTH_CAL_RATES = {
    "💪 Bicep-Tricep Curl": 3.5,   # upper isolation, moderate intensity: 2.7–3.9 kcal/min
    "🦵 Leg": 6.0,                 # lower machine (leg press): 5.0–7.3 kcal/min
    "🏋️ Strength- Others": 5.5,    # compound mixed: 5.5–6.0 kcal/min
}

SKIP_KEYBOARD = InlineKeyboardMarkup([[InlineKeyboardButton("⏭️ Skip", callback_data="skip")]])

DEFAULT_KEYBOARD = ReplyKeyboardMarkup(
    [["Log Workout", "⚖️ Weight"], ["🛌 Rest Day", "⚙️ Settings"]], resize_keyboard=True, one_time_keyboard=False
)


def _authorized(update: Update) -> bool:
    uid = update.effective_user.id if update.effective_user else 0
    return uid == config.USER_ID


def _today() -> str:
    return datetime.date.today().isoformat()


def compute_weight_math() -> Optional[str]:
    inputs = db.get_weight_math_inputs()
    if not inputs:
        return None

    profile = inputs["profile"]
    goal = inputs["goal"]
    target_date = inputs["target_date"]
    current_weight = inputs["current_weight"]
    metrics = inputs["metrics"]

    w = current_weight
    h_m = profile["height_cm"] / 100.0
    a = profile["age"]
    gender = profile["gender"]
    pal = profile["pal"]

    if gender == "male":
        bmr = (9.65 * w) + (573 * h_m) - (5.08 * a) + 260
    else:
        bmr = (10 * w) + (6.25 * profile["height_cm"]) - (5 * a) - 161

    tdee = bmr * pal

    today = datetime.date.today()
    days_remaining = None
    if target_date:
        try:
            target = datetime.date.fromisoformat(target_date)
            days_remaining = (target - today).days
        except (ValueError, TypeError):
            pass

    # Actual trend from 7-day MA regression
    weights = [m["weight_kg"] for m in metrics if m["weight_kg"]]
    actual_trend_kg_week = 0.0
    if len(weights) >= 7:
        ma = []
        for i in range(len(weights)):
            start = max(0, i - 6)
            slice_vals = weights[start:i + 1]
            ma.append(sum(slice_vals) / len(slice_vals))
        n = min(30, len(ma))
        x = list(range(n))
        y = ma[-n:]
        sum_x = sum(x)
        sum_y = sum(y)
        sum_xy = sum(xi * yi for xi, yi in zip(x, y))
        sum_x2 = sum(xi * xi for xi in x)
        denom = n * sum_x2 - sum_x * sum_x
        if denom != 0:
            slope = (n * sum_xy - sum_x * sum_y) / denom
            actual_trend_kg_week = slope * 7

    lines = ["⚖️ *Weight Math*"]
    lines.append(f"Current: *{round(current_weight, 1)}* kg → Goal: *{round(goal, 1)}* kg")
    lines.append(f"TDEE: *{round(tdee)}* kcal/day (BMR {round(bmr)}, PAL {pal})")

    if days_remaining is not None and days_remaining > 0:
        total_kg = current_weight - goal
        total_deficit = total_kg * 7700 * 1.12
        daily_deficit = total_deficit / days_remaining
        eat_target = tdee - daily_deficit
        required_trend = total_kg / (days_remaining / 7)

        lines.append(f"Days left: *{days_remaining}* (target: {target_date})")
        lines.append(f"Required rate: *{required_trend:.2f}* kg/week")

        # Primary: exercise burn target (assuming eating at TDEE)
        lines.append(f"")
        lines.append(f"🔥 *Exercise target: {round(daily_deficit)} kcal/day*")
        lines.append(f"   (if eating ~{round(tdee)} kcal/day)")

        if eat_target > 0:
            lines.append(f"🍽️ OR eat ~*{round(eat_target)}* kcal/day")
            lines.append(f"   (with your current exercise)")

        # Trend gap analysis (only meaningful with 2+ data points)
        if len(weights) >= 2:
            lines.append(f"")
            lines.append(f"Actual trend: *{actual_trend_kg_week:.2f}* kg/week")
            gap = required_trend - actual_trend_kg_week
            if gap > 0.05:
                extra_kcal = gap * 7700 / 7
                lines.append(f"🎯 Gap: +*{gap:.2f}* kg/week → burn *{round(extra_kcal)}* more kcal today")
            elif gap < -0.05:
                lines.append(f"✅ Ahead by *{abs(gap):.2f}* kg/week")
            else:
                lines.append("✅ On track")

        if eat_target < 1200:
            lines.append(f"⚠️ Eat target (*{round(eat_target)}* kcal) below safe minimum. Extend timeline.")
    elif days_remaining is not None and days_remaining <= 0:
        lines.append("⚠️ Target date has passed. Update with /target")
    else:
        lines.append("Set target date: /target 80 2026-08-01")

    return "\n".join(lines)


def fmt_report(today_rows, avg7) -> str:
    streak = db.get_streak()
    lines = []
    if streak > 1:
        lines.append(f"🔥 *{streak} day streak*")
    if db.is_rest_day(_today()):
        lines.append("🛌 *Rest Day*")
    lines.append("📋 *Today Total*")
    total_cal = sum(w["calories"] or 0 for w in today_rows)
    total_min = sum(w["duration_min"] or 0 for w in today_rows)
    if not today_rows:
        lines.append("No workouts logged yet.")
    else:
        for i, w in enumerate(today_rows, 1):
            name = w["machine"]
            lines.append(f"  • {name} — {w['calories']} kcal ({round(w['duration_min'])} min)")
    lines.append("")
    lines.append("📊 *Dashboard Now*")
    lines.append(f"  • Today calories burned: *{round(total_cal, 1)}* kcal")
    lines.append(f"  • Today workout time: *{round(total_min)}* min")
    lines.append(f"  • 7-day avg calories/day: *{avg7['avg_cal']}* kcal/day")
    lines.append(f"  • 7-day avg workout time/day: *{round(avg7['avg_min'])}* min/day")
    need = round(config.DAILY_GOAL_KCAL - total_cal, 1)
    if need > 0:
        lines.append(f"  • Need *{need}* more kcal to hit {config.DAILY_GOAL_KCAL}")
    else:
        lines.append(f"  • 🎯 Daily goal *{config.DAILY_GOAL_KCAL}* kcal reached!")
    return "\n".join(lines)


def _fix_dist(val: float, unit: str) -> float:
    """Try to fix missing decimal point in distance readings from OCR."""
    if val <= 50:
        return val
    s = str(int(val))
    for i in range(1, len(s)):
        try:
            c = float(s[:i] + "." + s[i:])
            if 0.1 <= c <= 20:
                return c
        except ValueError:
            pass
    return val


def _center(bbox):
    xs = [p[0] for p in bbox]
    ys = [p[1] for p in bbox]
    return sum(xs) / len(xs), sum(ys) / len(ys)


def _parse_spatial(raw, img_size):
    result = {"duration_min": None, "calories": None, "distance": None}
    if not raw or not img_size:
        return result
    width, _ = img_size
    threshold_x = width * 0.12

    items = []
    for r in raw:
        bbox, text, conf = r
        cx, cy = _center(bbox)
        items.append({"text": str(text), "cx": cx, "cy": cy, "conf": conf})

    labels = []
    for it in items:
        t = it["text"].lower()
        if re.search(r"calories|cal\b", t):
            labels.append(("calories", it))
        elif re.search(r"distance|dist\b", t):
            labels.append(("distance", it))
        elif re.search(r"time.*elapsed|elapsed.*time|time\b", t):
            labels.append(("time", it))

    values = []
    for it in items:
        t = it["text"]
        # Time with colon
        m = re.match(r"^(\d{1,2}):(\d{2})$", t)
        if m:
            val = int(m.group(1)) + int(m.group(2)) / 60
            if 1 <= val <= 300:
                values.append(("time", val, it))
        # Smooshed time 3-4 digits
        m = re.match(r"^(\d{3,4})$", t)
        if m:
            num = m.group(1)
            mm, ss = int(num[:2]), int(num[2:])
            if 0 <= mm <= 180 and 0 <= ss <= 59:
                val = mm + ss / 60
                if 1 <= val <= 300:
                    values.append(("time", val, it))
        # Pure integer (calories or distance)
        m = re.match(r"^(\d+)$", t)
        if m:
            val_int = int(m.group(1))
            if 50 <= val_int <= 2000:
                values.append(("calories", val_int, it))
            fixed = _fix_dist(float(val_int), "")
            if 0.1 <= fixed <= 50:
                values.append(("distance", fixed, it))
        # Float distance
        m = re.match(r"^(\d+\.\d+)$", t)
        if m:
            val = float(m.group(1))
            if 0.1 <= val <= 50:
                values.append(("distance", val, it))

    for label_type, lbl in labels:
        best = None
        best_score = None
        for vtype, val, vit in values:
            if vit["cy"] >= lbl["cy"]:
                continue
            dx = abs(vit["cx"] - lbl["cx"])
            if dx > threshold_x:
                continue
            score = dx
            if vtype == label_type:
                score -= 10000
            if best is None or score < best_score:
                best = (val, vtype)
                best_score = score

        if best:
            val, vtype = best
            if label_type == "time" and result["duration_min"] is None:
                result["duration_min"] = float(val)
            elif label_type == "calories" and result["calories"] is None:
                result["calories"] = int(val)
            elif label_type == "distance" and result["distance"] is None:
                result["distance"] = float(val)

    return result


def parse_ocr_text(text, raw=None, img_size=None) -> dict:
    """Parse OCR output from tesseract (str) or easyocr (list[str] or raw tuples)."""
    result: dict = {"duration_min": None, "calories": None, "distance": None}
    if not text:
        return result

    if raw is not None and img_size is not None:
        spatial = _parse_spatial(raw, img_size)
        if any(spatial.values()):
            return spatial

    if isinstance(text, str):
        texts = text.splitlines()
        full_text = text.lower()
    else:
        texts = text
        full_text = " ".join(str(t) for t in texts).lower()

    # DURATION — global patterns first
    m = re.search(r"time\s*(?:elapsed\s*)?(\d{1,2}):(\d{2})(?::(\d{2}))?", full_text)
    if not m:
        m = re.search(r"elapsed\s*(\d{1,2}):(\d{2})(?::(\d{2}))?", full_text)
    if m:
        groups = m.groups()
        if groups[2]:
            val = int(groups[0]) * 60 + int(groups[1]) + int(groups[2]) / 60
        else:
            val = int(groups[0]) + int(groups[1]) / 60
        if 1 <= val <= 300:
            result["duration_min"] = val

    if not result["duration_min"]:
        m = re.search(r"(\d{1,2}):(\d{2})", full_text)
        if m:
            val = int(m.group(1)) + int(m.group(2)) / 60
            if 1 <= val <= 300:
                result["duration_min"] = val

    # Smooshed 4-digit time anywhere → MMSS
    if not result["duration_min"]:
        for t in texts:
            m = re.match(r"^(\d{3,4})$", str(t))
            if m:
                num = m.group(1)
                mm, ss = int(num[:2]), int(num[2:])
                if 0 <= mm <= 180 and 0 <= ss <= 59:
                    val = mm + ss / 60
                    if 1 <= val <= 300:
                        result["duration_min"] = val
                        break

    # CALORIES — bidirectional scan around label
    for i, t in enumerate(texts):
        if re.search(r"calories|cal\b", str(t), re.I):
            for j in range(max(0, i - 5), min(len(texts), i + 4)):
                if j == i:
                    continue
                m = re.search(r"(\d{2,4})", str(texts[j]))
                if m:
                    val = int(m.group(1))
                    if 50 <= val <= 2000:
                        result["calories"] = val
                        break
            if result["calories"]:
                break

    # DISTANCE — bidirectional scan around label
    for i, t in enumerate(texts):
        if re.search(r"distance|dist\b", str(t), re.I):
            for j in range(max(0, i - 5), min(len(texts), i + 4)):
                if j == i:
                    continue
                item = str(texts[j])
                m = re.search(r"(\d+\.\d+)", item)
                if m:
                    val = float(m.group(1))
                    if 0.1 <= val <= 50:
                        result["distance"] = val
                        break
                m = re.search(r"(\d+\.?\d*)(km|mi)", item, re.I)
                if m:
                    val = float(m.group(1))
                    val = _fix_dist(val, m.group(2).lower())
                    if 0.1 <= val <= 50:
                        result["distance"] = val
                        break
                # Try plain integer with _fix_dist
                m = re.search(r"^(\d+)$", item)
                if m:
                    val = _fix_dist(float(m.group(1)), "")
                    if 0.1 <= val <= 50:
                        result["distance"] = val
                        break
            if result["distance"]:
                break

    # Fallback: any number+km/mi anywhere
    if not result["distance"]:
        m = re.search(r"(\d+\.?\d*)\s*(km|mi)", full_text)
        if m:
            val = float(m.group(1))
            val = _fix_dist(val, m.group(2).lower())
            if 0.1 <= val <= 50:
                result["distance"] = val

    return result


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        await update.message.reply_text("Unauthorized.")
        return
    today_rows = db.get_workouts_for_date(_today())
    avg7 = db.get_7day_avg()
    await update.message.reply_text(
        fmt_report(today_rows, avg7), parse_mode="Markdown", reply_markup=DEFAULT_KEYBOARD
    )


async def log_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not _authorized(update):
        return ConversationHandler.END
    context.user_data.clear()
    keyboard = [
        [InlineKeyboardButton(m, callback_data=m) for m in row]
        for row in MACHINE_OPTIONS
    ]
    await update.message.reply_text(
        "Select machine:", reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return TYPE


async def photo_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not _authorized(update):
        return ConversationHandler.END
    # Guard: if already mid-workout log, don't restart
    if "machine" in context.user_data:
        await update.message.reply_text(
            "You're already logging a workout. Finish or /cancel first."
        )
        return ConversationHandler.END
    context.user_data.clear()
    photo = update.message.photo[-1]
    try:
        file = await photo.get_file()
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fp = file.file_path or ""
        ext = os.path.splitext(fp)[1] or ".jpg"
        path = os.path.join(config.PHOTO_DIR, f"workout_{ts}{ext}")
        await file.download_to_drive(path)
    except Exception as e:
        logger.error("Photo download failed: %s", e)
        await update.message.reply_text("❌ Failed to download photo. Try again or use /log.")
        return ConversationHandler.END
    context.user_data["photo_path"] = path

    def _ocr_easy(path: str) -> dict:
        if _EASYOCR_READER is None:
            return {}
        try:
            from PIL import Image
            import numpy as np
            img = Image.open(path)
            w, h = img.size
            # Auto-crop phone screenshots
            if h > w * 1.5:
                img = img.crop((int(w*0.05), int(h*0.28), int(w*0.95), int(h*0.72)))
            raw = _EASYOCR_READER.readtext(np.array(img))
            texts = [r[1] for r in raw]
            result = parse_ocr_text(texts, raw=raw, img_size=img.size)
            logger.info("EasyOCR result: %s", result)
            return result
        except Exception as e:
            logger.info("EasyOCR failed: %s", e)
            return {}

    def _ocr_tesseract(path: str) -> dict:
        try:
            import pytesseract
            from PIL import Image
            pytesseract.pytesseract.tesseract_cmd = "/opt/homebrew/bin/tesseract"
            img = Image.open(path)
            w, h = img.size
            # Auto-crop phone screenshots
            if h > w * 1.5:
                img = img.crop((int(w*0.05), int(h*0.28), int(w*0.95), int(h*0.72)))
            text = pytesseract.image_to_string(img)
            result = parse_ocr_text(text)
            logger.info("Tesseract result: %s", result)
            return result
        except Exception as e:
            logger.info("Tesseract failed: %s", e)
            return {}

    # Run LLM and EasyOCR in parallel threads.
    # LLM is higher quality; EasyOCR is the instant fallback if LLM fails.
    easy_task = asyncio.create_task(asyncio.to_thread(_ocr_easy, path))
    llm_task = asyncio.create_task(asyncio.to_thread(llm_ocr.llm_ocr, path))

    ocr_result = {}
    llm_result = await llm_task
    if isinstance(llm_result, dict) and llm_ocr._is_valid(llm_result):
        ocr_result = llm_result
        logger.info("OCR winner: LLM → %s", ocr_result)
        easy_task.cancel()
    else:
        logger.warning("LLM OCR failed or invalid, trying EasyOCR fallback")
        try:
            easy_result = await easy_task
            if isinstance(easy_result, dict) and any(easy_result.values()):
                ocr_result = easy_result
                logger.info("OCR winner: EasyOCR fallback → %s", ocr_result)
        except asyncio.CancelledError:
            pass

    # Final fallback: Tesseract
    if not ocr_result:
        ocr_result = await asyncio.to_thread(_ocr_tesseract, path)
        if ocr_result:
            logger.info("OCR winner: Tesseract → %s", ocr_result)

    context.user_data["ocr"] = ocr_result

    pre = []
    if ocr_result.get("duration_min"):
        pre.append(f"Duration: {round(ocr_result['duration_min'])} min")
    if ocr_result.get("calories"):
        pre.append(f"Calories: {ocr_result['calories']} kcal")
    if ocr_result.get("distance"):
        pre.append(f"Distance: {ocr_result['distance']} km")

    msg = "Photo received."
    if pre:
        msg += " I read:\n" + "\n".join(pre) + "\n\nSelect machine:"
    else:
        msg += " (Couldn't read numbers from photo — you'll enter them manually.)\n\nSelect machine:"

    keyboard = [
        [InlineKeyboardButton(m, callback_data=m) for m in row]
        for row in MACHINE_OPTIONS
    ]
    await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(keyboard))
    return TYPE


async def type_selected(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    machine = query.data
    context.user_data["machine"] = machine
    context.user_data["workout_type"] = TYPE_MAP.get(machine, "Cardio")

    ocr = context.user_data.get("ocr", {})
    if ocr.get("duration_min"):
        context.user_data["duration_min"] = ocr["duration_min"]
        await query.edit_message_text(
            f"Machine: {machine}\nOCR duration: {round(ocr['duration_min'])} min\n\nReply with duration in minutes, or tap Skip to keep."
        )
        await query.message.reply_text("Duration:", reply_markup=SKIP_KEYBOARD)
    else:
        await query.edit_message_text(f"Machine: {machine}\n\nReply with duration in minutes:")
    return DURATION


async def duration_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "/cancel":
        return await cancel(update, context)
    try:
        context.user_data["duration_min"] = float(text)
    except ValueError:
        await update.message.reply_text("Send a number for duration (minutes), or /cancel:")
        return DURATION

    return await _ask_calories(update, context)


async def duration_skip(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if "duration_min" not in context.user_data:
        await query.edit_message_text("No OCR duration to skip. Reply with duration in minutes:")
        return DURATION
    await query.edit_message_text(f"Duration: {round(context.user_data['duration_min'])} min ✅")
    return await _ask_calories(update, context)


async def _ask_calories(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    msg_obj = update.message or update.callback_query.message
    d = context.user_data
    if "duration_min" not in d:
        await msg_obj.reply_text("Something went wrong. Start over with /log")
        d.clear()
        return ConversationHandler.END
    ocr = d.get("ocr", {})
    machine = d.get("machine", "")
    if ocr.get("calories"):
        d["calories"] = ocr["calories"]
        msg = f"Duration: {round(d['duration_min'])} min\nOCR calories: {ocr['calories']}\n\nReply with calories, or tap Skip to keep."
    elif machine in STRENGTH_CAL_RATES:
        rate = STRENGTH_CAL_RATES[machine]
        est = round(d['duration_min'] * rate)
        msg = f"Reply with calories, or tap Skip to auto-estimate (~{est} kcal):"
    else:
        msg = "Reply with calories shown on machine:"
    await msg_obj.reply_text(msg, reply_markup=SKIP_KEYBOARD)
    return CALORIES


async def calories_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "/cancel":
        return await cancel(update, context)
    try:
        context.user_data["calories"] = float(text)
    except ValueError:
        await update.message.reply_text("Send a number for calories, or /cancel:")
        return CALORIES

    return await _ask_distance(update, context)


async def calories_skip(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    machine = context.user_data.get("machine", "")
    if "calories" not in context.user_data:
        if machine in STRENGTH_CAL_RATES and "duration_min" in context.user_data:
            estimated = round(context.user_data["duration_min"] * STRENGTH_CAL_RATES[machine])
            context.user_data["calories"] = estimated
            await query.edit_message_text(f"Calories: {estimated} kcal (auto-estimated) ✅")
            return await _ask_distance(update, context)
        await query.edit_message_text("No calories to skip. Reply with calories:")
        return CALORIES
    await query.edit_message_text(f"Calories: {context.user_data['calories']} kcal ✅")
    return await _ask_distance(update, context)


async def _ask_distance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    msg_obj = update.message or update.callback_query.message
    d = context.user_data
    if "calories" not in d:
        await msg_obj.reply_text("Something went wrong. Start over with /log")
        d.clear()
        return ConversationHandler.END
    ocr = d.get("ocr", {})
    if ocr.get("distance"):
        d["distance"] = ocr["distance"]
        msg = f"Calories: {d['calories']}\nOCR distance: {ocr['distance']} km\n\nReply with distance (km), or tap Skip to keep."
    else:
        msg = "Reply with distance (km), or tap Skip if not applicable:"
    await msg_obj.reply_text(msg, reply_markup=SKIP_KEYBOARD)
    return DISTANCE


async def distance_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "/cancel":
        return await cancel(update, context)

    miles_match = re.search(r"([\d.]+)\s*(?:miles?|mi)", text, re.IGNORECASE)
    if miles_match:
        miles = float(miles_match.group(1))
        km = round(miles * 1.60934, 2)
        context.user_data["distance"] = km
        await update.message.reply_text(f"Converted {miles} mi → {km} km")
    else:
        try:
            context.user_data["distance"] = float(text)
        except ValueError:
            await update.message.reply_text("Send a number for distance, or /cancel:")
            return DISTANCE

    return await _ask_confirm(update, context)


async def distance_skip(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    await query.edit_message_text("Distance skipped ✅")
    return await _ask_confirm(update, context)


async def _ask_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    d = context.user_data
    required = ["machine", "duration_min", "calories"]
    missing = [k for k in required if k not in d]
    if missing:
        msg_obj = update.message or update.callback_query.message
        await msg_obj.reply_text(
            f"Missing data ({', '.join(missing)}). Start over with /log"
        )
        d.clear()
        return ConversationHandler.END
    summary = (
        f"Confirm log:\n"
        f"  • {d['machine']} ({d['workout_type']})\n"
        f"  • {round(d['duration_min'])} min\n"
        f"  • {d['calories']} kcal"
    )
    if "distance" in d:
        summary += f"\n  • {d['distance']} km"

    keyboard = [
        [
            InlineKeyboardButton("✅ Confirm", callback_data="confirm"),
            InlineKeyboardButton("❌ Cancel", callback_data="cancel"),
        ]
    ]
    msg_obj = update.message or update.callback_query.message
    await msg_obj.reply_text(summary, reply_markup=InlineKeyboardMarkup(keyboard))
    return CONFIRM


async def confirm_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        await query.edit_message_text("Cancelled.")
        context.user_data.clear()
        return ConversationHandler.END

    if context.user_data.get("_confirmed"):
        await query.edit_message_text("Already saved.")
        return ConversationHandler.END

    d = context.user_data
    required = ["machine", "duration_min", "calories"]
    missing = [k for k in required if k not in d]
    if missing:
        await query.edit_message_text(f"Missing: {', '.join(missing)}. Start over with /log")
        context.user_data.clear()
        return ConversationHandler.END

    machine = d["machine"]
    calories = d["calories"]

    try:
        prs = db.get_machine_prs()
        previous_best = prs.get(machine)

        context.user_data["_confirmed"] = True
        wid = db.add_workout(
            date=_today(),
            workout_type=d.get("workout_type", "Cardio"),
            machine=machine,
            duration_min=d["duration_min"],
            calories=calories,
            level=d.get("level"),
            distance=d.get("distance"),
            notes=d.get("notes"),
            photo_path=d.get("photo_path"),
        )
    except Exception as e:
        logger.error("Confirm handler error: %s", e)
        await query.edit_message_text(f"❌ Error saving workout: {e}\nTry /log again.")
        context.user_data.clear()
        return ConversationHandler.END

    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", f"Auto-export workouts after log {wid}"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)

    today_rows = db.get_workouts_for_date(_today())
    avg7 = db.get_7day_avg()
    await query.edit_message_text(fmt_report(today_rows, avg7), parse_mode="Markdown")

    if previous_best is None or calories > previous_best:
        msg = f"🎉 New PR on {machine}: {calories} kcal"
        if previous_best:
            msg += f" (previous best: {previous_best})"
        await context.bot.send_message(chat_id=config.USER_ID, text=msg)

    # Weight math nudge after workout
    math_msg = compute_weight_math()
    if math_msg:
        await context.bot.send_message(chat_id=config.USER_ID, text=math_msg, parse_mode="Markdown")

    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not _authorized(update):
        return ConversationHandler.END
    await update.message.reply_text("Cancelled.")
    context.user_data.clear()
    return ConversationHandler.END


async def export_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        await update.message.reply_text("❌ openpyxl not installed.")
        return

    xlsx_path = "/Users/billkim/gym-tracker/workout_tracker.xlsx"
    wb = None
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workout Log"
        headers = ["Date", "Day", "Type", "Machine", "Duration (min)", "Calories", "Adjusted Calories",
                   "Level", "Distance", "Floors/Steps", "Weight/Load", "Sets/Reps", "Notes", "Week Start", "Month"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        wb.save(xlsx_path)
        wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Workout Log"]

    row = 2
    while ws.cell(row=row, column=1).value is not None:
        row += 1

    workouts = db.get_recent_workouts(30)
    existing_dates_machines = set()
    for r in range(2, row):
        d = ws.cell(row=r, column=1).value
        m = ws.cell(row=r, column=4).value
        if d and m:
            if hasattr(d, "isoformat"):
                d = d.isoformat()
            existing_dates_machines.add((str(d)[:10], str(m)))

    added = 0
    for w in workouts:
        key = (w["date"], w["machine"])
        if key in existing_dates_machines:
            continue
        ws.cell(row=row, column=1, value=w["date"])
        ws.cell(row=row, column=2, value=f"=IF(A{row}=\"\",\"\",TEXT(A{row},\"ddd\"))")
        ws.cell(row=row, column=3, value=w["type"])
        ws.cell(row=row, column=4, value=w["machine"])
        ws.cell(row=row, column=5, value=w["duration_min"])
        ws.cell(row=row, column=6, value=w["calories"])
        ws.cell(row=row, column=7, value=f"=IF(F{row}=\"\",\"\",F{row}*Settings!\$B\$3)")
        ws.cell(row=row, column=8, value=w["level"])
        ws.cell(row=row, column=9, value=w["distance"])
        ws.cell(row=row, column=10, value=w["floors_steps"])
        ws.cell(row=row, column=11, value=w["weight_load"])
        ws.cell(row=row, column=12, value=w["sets_reps"])
        ws.cell(row=row, column=13, value=w["notes"])
        ws.cell(row=row, column=14, value=f"=IF(A{row}=\"\",\"\",A{row}-WEEKDAY(A{row},2)+1)")
        ws.cell(row=row, column=15, value=f"=IF(A{row}=\"\",\"\",TEXT(A{row},\"yyyy-mm\"))")
        row += 1
        added += 1

    try:
        wb.save(xlsx_path)
    except Exception as e:
        await update.message.reply_text(f"❌ Failed to save xlsx: {e}")
        return
    await update.message.reply_text(f"Exported {added} new workouts to xlsx.")


async def weight_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    # Cancel any active conversation
    context.user_data.clear()
    text = update.message.text.replace("/weight", "").strip()
    if not text:
        await update.message.reply_text("Usage: /weight 87.5")
        return
    try:
        w = float(text)
    except ValueError:
        await update.message.reply_text("Usage: /weight 87.5")
        return
    if w <= 0:
        await update.message.reply_text("Weight must be positive.")
        return
    await _log_weight(update, context, w)


async def weight_button_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not _authorized(update):
        return ConversationHandler.END
    context.user_data.clear()
    await update.message.reply_text("Send your weight in kg (e.g., 87.5):")
    return WAITING_WEIGHT


async def weight_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "/cancel":
        await update.message.reply_text("Cancelled.")
        context.user_data.clear()
        return ConversationHandler.END
    # Allow "/weight 86.1" as a reply, not just "86.1"
    text = text.replace("/weight", "").strip()
    try:
        w = float(text)
    except ValueError:
        await update.message.reply_text("Send a number for weight in kg, or /cancel:")
        return WAITING_WEIGHT
    if w <= 0:
        await update.message.reply_text("Weight must be positive. Try again or /cancel:")
        return WAITING_WEIGHT
    await _log_weight(update, context, w)
    context.user_data.clear()
    return ConversationHandler.END


async def _log_weight(update: Update, context: ContextTypes.DEFAULT_TYPE, w: float) -> None:
    db.add_body_metric(_today(), weight_kg=w)
    reply = f"✅ Logged weight: {w} kg"
    proj = db.weight_projection()
    if proj:
        reply += f"\n📊 7-day avg: {proj['current_ma']} kg"
        reply += f"\n📉 Trend: {proj['slope_kg_week']:+} kg/week"
        if proj['weeks_to_goal'] is not None:
            reply += f"\n🎯 Goal {proj['goal']} kg: ~{proj['weeks_to_goal']} weeks"
        if proj['message']:
            reply += f"\n{proj['message']}"
    await update.message.reply_text(reply)

    # Append full calorie math if profile + target exist
    math_msg = compute_weight_math()
    if math_msg:
        await update.message.reply_text(math_msg, parse_mode="Markdown")

    # Auto-export and push so GitHub Pages dashboard stays current
    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", f"Auto-export: weight {w} kg"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)


async def goal_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    text = update.message.text.replace("/goal", "").strip()
    if not text:
        current = db.get_goal_weight()
        if current:
            await update.message.reply_text(f"Current goal: {current} kg\nTo change: /goal 80")
        else:
            await update.message.reply_text("No goal set. Use /goal 80")
        return
    try:
        w = float(text)
    except ValueError:
        await update.message.reply_text("Usage: /goal 80")
        return
    if w <= 0:
        await update.message.reply_text("Goal must be positive.")
        return
    db.set_goal_weight(w)
    await update.message.reply_text(f"Goal set: {w} kg")

    # Auto-export and push so GitHub Pages dashboard picks up the new goal
    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", f"Auto-export: set goal {w} kg"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)


async def me_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    text = update.message.text.replace("/me", "").strip()
    if not text:
        profile = db.get_profile()
        if profile:
            await update.message.reply_text(
                f"Profile:\nHeight: {profile['height_cm']} cm\nAge: {profile['age']}\nGender: {profile['gender']}\nPAL: {profile['pal']}\n\nTo update: /me height 170 age 30 gender male"
            )
        else:
            await update.message.reply_text("No profile. Set it: /me height 170 age 30 gender male")
        return
    # Parse loose key-value pairs
    parts = text.lower().split()
    data = {}
    i = 0
    while i < len(parts):
        if parts[i] in ("height", "h") and i + 1 < len(parts):
            data["height"] = parts[i + 1]
            i += 2
        elif parts[i] in ("age", "a") and i + 1 < len(parts):
            data["age"] = parts[i + 1]
            i += 2
        elif parts[i] in ("gender", "g") and i + 1 < len(parts):
            data["gender"] = parts[i + 1]
            i += 2
        elif parts[i] in ("pal", "p") and i + 1 < len(parts):
            data["pal"] = parts[i + 1]
            i += 2
        else:
            i += 1
    if not data:
        await update.message.reply_text("Usage: /me height 170 age 30 gender male")
        return
    profile = db.get_profile() or {}
    try:
        height = float(data.get("height", profile.get("height_cm", 170)))
        age = int(data.get("age", profile.get("age", 30)))
        gender = data.get("gender", profile.get("gender", "male"))
        pal = float(data.get("pal", profile.get("pal", 1.4)))
    except (ValueError, TypeError):
        await update.message.reply_text("Invalid values. Example: /me height 170 age 30 gender male")
        return
    if height <= 0 or age <= 0 or pal < 1.0 or pal > 2.5:
        await update.message.reply_text("Invalid range: height/age must be positive, PAL 1.0–2.5.")
        return
    db.set_profile(height, age, gender, pal)
    await update.message.reply_text(f"Profile saved: {height} cm, {age} yr, {gender}, PAL {pal}")

    # Auto-export so dashboard reflects profile changes
    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", f"Auto-export: profile update"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)


async def target_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    text = update.message.text.replace("/target", "").strip()
    if not text:
        current_goal = db.get_goal_weight()
        current_target = db.get_target_date()
        lines = []
        if current_goal:
            lines.append(f"Goal weight: {current_goal} kg")
        if current_target:
            lines.append(f"Target date: {current_target}")
        if lines:
            await update.message.reply_text("\n".join(lines) + "\n\nTo change: /target 80 2026-08-01")
        else:
            await update.message.reply_text("No target set. Use /target 80 2026-08-01")
        return
    parts = text.split()
    if len(parts) < 2:
        await update.message.reply_text("Usage: /target 80 2026-08-01")
        return
    try:
        goal = float(parts[0])
        target_date = parts[1]
        # Validate date
        datetime.date.fromisoformat(target_date)
    except (ValueError, TypeError):
        await update.message.reply_text("Usage: /target 80 2026-08-01")
        return
    db.set_goal_weight(goal)
    db.set_target_date(target_date)
    await update.message.reply_text(f"Target set: {goal} kg by {target_date}")
    # Auto-export
    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", f"Auto-export: target {goal} kg by {target_date}"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)


async def math_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    msg = compute_weight_math()
    if msg:
        await update.message.reply_text(msg, parse_mode="Markdown")
    else:
        await update.message.reply_text(
            "Need profile + goal + weight data.\n"
            "1. /me height 170 age 30 gender male\n"
            "2. /target 80 2026-08-01\n"
            "3. Log weight with /weight"
        )


# ---- Button-based Settings ----

async def settings_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 Weight Math", callback_data="math"),
         InlineKeyboardButton("🎯 Set Target", callback_data="set_target")],
        [InlineKeyboardButton("👤 Edit Profile", callback_data="set_profile")],
    ])
    await update.message.reply_text("What do you want to do?", reply_markup=kb)


async def settings_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "math":
        msg = compute_weight_math()
        if msg:
            await query.edit_message_text(msg, parse_mode="Markdown")
        else:
            await query.edit_message_text(
                "Need profile + goal + weight. Use \u2699\ufe0f Settings → Edit Profile / Set Target first."
            )
    elif data == "set_target":
        await target_weight_start(update, context)
    elif data == "set_profile":
        await profile_start(update, context)


# Target conversation (weight → date)
TARGET_WEIGHT, TARGET_DATE = range(2)

async def target_weight_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    goal = db.get_goal_weight()
    target = db.get_target_date()
    header = ""
    if goal and target:
        header = f"Current: {goal} kg by {target}\n"
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("75", callback_data="tw_75"),
         InlineKeyboardButton("78", callback_data="tw_78"),
         InlineKeyboardButton("80", callback_data="tw_80")],
        [InlineKeyboardButton("82", callback_data="tw_82"),
         InlineKeyboardButton("85", callback_data="tw_85"),
         InlineKeyboardButton("Type manually", callback_data="tw_manual")],
        [InlineKeyboardButton("❌ Cancel", callback_data="tw_cancel")],
    ])
    await query.edit_message_text(f"{header}Pick goal weight (kg):", reply_markup=kb)
    return TARGET_WEIGHT


async def target_weight_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data
    if data.startswith("tw_"):
        val = data.replace("tw_", "")
        if val == "manual":
            await query.edit_message_text("Send goal weight in kg (e.g. 80):")
            return TARGET_WEIGHT
        if val == "cancel":
            await query.edit_message_text("Cancelled.")
            context.user_data.clear()
            return ConversationHandler.END
        try:
            w = float(val)
        except ValueError:
            await query.edit_message_text("Invalid. Send goal weight in kg:")
            return TARGET_WEIGHT
    else:
        await query.edit_message_text("Send goal weight in kg:")
        return TARGET_WEIGHT
    context.user_data["_target_weight"] = w
    return await _ask_target_date(query, context)


async def target_weight_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    try:
        w = float(text)
    except ValueError:
        await update.message.reply_text("Send a number for goal weight in kg:")
        return TARGET_WEIGHT
    if w <= 0:
        await update.message.reply_text("Must be positive. Try again:")
        return TARGET_WEIGHT
    context.user_data["_target_weight"] = w
    return await _ask_target_date_msg(update, context)


async def _ask_target_date(query, context) -> int:
    today = datetime.date.today()
    dates = [
        (today + datetime.timedelta(days=14), "2 weeks"),
        (today + datetime.timedelta(days=30), "1 month"),
        (today + datetime.timedelta(days=60), "2 months"),
        (today + datetime.timedelta(days=90), "3 months"),
    ]
    buttons = [[InlineKeyboardButton(label, callback_data=f"td_{d.isoformat()}")]
               for d, label in dates]
    buttons.append([InlineKeyboardButton("Type manually", callback_data="td_manual")])
    buttons.append([InlineKeyboardButton("❌ Cancel", callback_data="td_cancel")])
    await query.edit_message_text("Pick target date:", reply_markup=InlineKeyboardMarkup(buttons))
    return TARGET_DATE


async def _ask_target_date_msg(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    today = datetime.date.today()
    dates = [
        (today + datetime.timedelta(days=14), "2 weeks"),
        (today + datetime.timedelta(days=30), "1 month"),
        (today + datetime.timedelta(days=60), "2 months"),
        (today + datetime.timedelta(days=90), "3 months"),
    ]
    buttons = [[InlineKeyboardButton(label, callback_data=f"td_{d.isoformat()}")]
               for d, label in dates]
    buttons.append([InlineKeyboardButton("Type manually", callback_data="td_manual")])
    buttons.append([InlineKeyboardButton("❌ Cancel", callback_data="td_cancel")])
    await update.message.reply_text("Pick target date:", reply_markup=InlineKeyboardMarkup(buttons))
    return TARGET_DATE


async def target_date_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data
    if data.startswith("td_"):
        val = data.replace("td_", "")
        if val == "manual":
            await query.edit_message_text("Send target date (YYYY-MM-DD):")
            return TARGET_DATE
        if val == "cancel":
            await query.edit_message_text("Cancelled.")
            context.user_data.clear()
            return ConversationHandler.END
        target_date = val
    else:
        await query.edit_message_text("Send target date (YYYY-MM-DD):")
        return TARGET_DATE
    return await _save_target(query, context, target_date)


async def target_date_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    try:
        datetime.date.fromisoformat(text)
    except ValueError:
        await update.message.reply_text("Invalid date. Send YYYY-MM-DD:")
        return TARGET_DATE
    return await _save_target_msg(update, context, text)


async def _save_target(query, context, target_date: str) -> int:
    w = context.user_data.get("_target_weight")
    if w is None:
        await query.edit_message_text("Something went wrong. Start over with ⚙️ Settings.")
        context.user_data.clear()
        return ConversationHandler.END
    db.set_goal_weight(w)
    db.set_target_date(target_date)
    await query.edit_message_text(f"🎯 Target saved: {w} kg by {target_date}")
    context.user_data.clear()
    # Auto-export
    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", f"Auto-export: target {w} kg by {target_date}"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)
    return ConversationHandler.END


async def _save_target_msg(update: Update, context: ContextTypes.DEFAULT_TYPE, target_date: str) -> int:
    w = context.user_data.get("_target_weight")
    if w is None:
        await update.message.reply_text("Something went wrong. Start over with ⚙️ Settings.")
        context.user_data.clear()
        return ConversationHandler.END
    db.set_goal_weight(w)
    db.set_target_date(target_date)
    await update.message.reply_text(f"🎯 Target saved: {w} kg by {target_date}", reply_markup=DEFAULT_KEYBOARD)
    context.user_data.clear()
    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", f"Auto-export: target {w} kg by {target_date}"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)
    return ConversationHandler.END


# Profile conversation (height → age → gender → PAL)
PROFILE_HEIGHT, PROFILE_AGE, PROFILE_GENDER, PROFILE_PAL = range(4)

async def profile_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    profile = db.get_profile()
    header = ""
    if profile:
        header = f"Current: {profile['height_cm']} cm, {profile['age']} yr, {profile['gender']}, PAL {profile['pal']}\n"
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("160", callback_data="ph_160"),
         InlineKeyboardButton("170", callback_data="ph_170"),
         InlineKeyboardButton("175", callback_data="ph_175"),
         InlineKeyboardButton("180", callback_data="ph_180")],
        [InlineKeyboardButton("Type manually", callback_data="ph_manual")],
        [InlineKeyboardButton("❌ Cancel", callback_data="ph_cancel")],
    ])
    await query.edit_message_text(f"{header}Pick height (cm):", reply_markup=kb)
    return PROFILE_HEIGHT


async def profile_height_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data
    if data.startswith("ph_"):
        val = data.replace("ph_", "")
        if val == "manual":
            await query.edit_message_text("Send height in cm:")
            return PROFILE_HEIGHT
        if val == "cancel":
            await query.edit_message_text("Cancelled.")
            context.user_data.clear()
            return ConversationHandler.END
        try:
            h = float(val)
        except ValueError:
            await query.edit_message_text("Invalid. Send height in cm:")
            return PROFILE_HEIGHT
    else:
        await query.edit_message_text("Send height in cm:")
        return PROFILE_HEIGHT
    context.user_data["_profile_height"] = h
    return await _ask_profile_age(query, context)


async def profile_height_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    try:
        h = float(text)
    except ValueError:
        await update.message.reply_text("Send a number for height in cm:")
        return PROFILE_HEIGHT
    if h <= 0:
        await update.message.reply_text("Must be positive. Try again:")
        return PROFILE_HEIGHT
    context.user_data["_profile_height"] = h
    return await _ask_profile_age_msg(update, context)


async def _ask_profile_age(query, context) -> int:
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("25", callback_data="pa_25"),
         InlineKeyboardButton("30", callback_data="pa_30"),
         InlineKeyboardButton("35", callback_data="pa_35"),
         InlineKeyboardButton("40", callback_data="pa_40")],
        [InlineKeyboardButton("Type manually", callback_data="pa_manual")],
        [InlineKeyboardButton("❌ Cancel", callback_data="pa_cancel")],
    ])
    await query.edit_message_text("Pick age:", reply_markup=kb)
    return PROFILE_AGE


async def _ask_profile_age_msg(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("25", callback_data="pa_25"),
         InlineKeyboardButton("30", callback_data="pa_30"),
         InlineKeyboardButton("35", callback_data="pa_35"),
         InlineKeyboardButton("40", callback_data="pa_40")],
        [InlineKeyboardButton("Type manually", callback_data="pa_manual")],
        [InlineKeyboardButton("❌ Cancel", callback_data="pa_cancel")],
    ])
    await update.message.reply_text("Pick age:", reply_markup=kb)
    return PROFILE_AGE


async def profile_age_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data
    if data.startswith("pa_"):
        val = data.replace("pa_", "")
        if val == "manual":
            await query.edit_message_text("Send age in years:")
            return PROFILE_AGE
        if val == "cancel":
            await query.edit_message_text("Cancelled.")
            context.user_data.clear()
            return ConversationHandler.END
        try:
            a = int(val)
        except ValueError:
            await query.edit_message_text("Invalid. Send age in years:")
            return PROFILE_AGE
    else:
        await query.edit_message_text("Send age in years:")
        return PROFILE_AGE
    context.user_data["_profile_age"] = a
    return await _ask_profile_gender(query, context)


async def profile_age_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    try:
        a = int(text)
    except ValueError:
        await update.message.reply_text("Send a whole number for age:")
        return PROFILE_AGE
    if a <= 0:
        await update.message.reply_text("Must be positive. Try again:")
        return PROFILE_AGE
    context.user_data["_profile_age"] = a
    return await _ask_profile_gender_msg(update, context)


async def _ask_profile_gender(query, context) -> int:
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Male", callback_data="pg_male"),
         InlineKeyboardButton("Female", callback_data="pg_female")],
        [InlineKeyboardButton("❌ Cancel", callback_data="pg_cancel")],
    ])
    await query.edit_message_text("Pick gender:", reply_markup=kb)
    return PROFILE_GENDER


async def _ask_profile_gender_msg(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Male", callback_data="pg_male"),
         InlineKeyboardButton("Female", callback_data="pg_female")],
        [InlineKeyboardButton("❌ Cancel", callback_data="pg_cancel")],
    ])
    await update.message.reply_text("Pick gender:", reply_markup=kb)
    return PROFILE_GENDER


async def profile_gender_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "pg_male":
        g = "male"
    elif data == "pg_female":
        g = "female"
    elif data == "pg_cancel":
        await query.edit_message_text("Cancelled.")
        context.user_data.clear()
        return ConversationHandler.END
    else:
        await query.edit_message_text("Pick Male or Female:")
        return PROFILE_GENDER
    context.user_data["_profile_gender"] = g
    return await _ask_profile_pal(query, context)


async def _ask_profile_pal(query, context) -> int:
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Desk Job 1.4", callback_data="pp_1.4")],
        [InlineKeyboardButton("Light 1.55", callback_data="pp_1.55")],
        [InlineKeyboardButton("Moderate 1.725", callback_data="pp_1.725")],
        [InlineKeyboardButton("Very Active 1.9", callback_data="pp_1.9")],
        [InlineKeyboardButton("❌ Cancel", callback_data="pp_cancel")],
    ])
    await query.edit_message_text("Pick activity level:", reply_markup=kb)
    return PROFILE_PAL


async def _ask_profile_pal_msg(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("Desk Job 1.4", callback_data="pp_1.4")],
        [InlineKeyboardButton("Light 1.55", callback_data="pp_1.55")],
        [InlineKeyboardButton("Moderate 1.725", callback_data="pp_1.725")],
        [InlineKeyboardButton("Very Active 1.9", callback_data="pp_1.9")],
        [InlineKeyboardButton("❌ Cancel", callback_data="pp_cancel")],
    ])
    await update.message.reply_text("Pick activity level:", reply_markup=kb)
    return PROFILE_PAL


async def profile_pal_received(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    data = query.data
    if data.startswith("pp_"):
        if data == "pp_cancel":
            await query.edit_message_text("Cancelled.")
            context.user_data.clear()
            return ConversationHandler.END
        try:
            pal = float(data.replace("pp_", ""))
        except ValueError:
            await query.edit_message_text("Pick an activity level:")
            return PROFILE_PAL
    else:
        await query.edit_message_text("Pick an activity level:")
        return PROFILE_PAL
    return await _save_profile(query, context, pal)


async def profile_pal_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    try:
        pal = float(text)
    except ValueError:
        await update.message.reply_text("Send a number for PAL (e.g. 1.4):")
        return PROFILE_PAL
    if pal < 1.0 or pal > 2.5:
        await update.message.reply_text("PAL should be 1.0–2.5. Try again:")
        return PROFILE_PAL
    return await _save_profile_msg(update, context, pal)


async def _save_profile(query, context, pal: float) -> int:
    d = context.user_data
    height = d.get("_profile_height")
    age = d.get("_profile_age")
    gender = d.get("_profile_gender")
    if None in (height, age, gender):
        await query.edit_message_text("Something went wrong. Start over with ⚙️ Settings.")
        context.user_data.clear()
        return ConversationHandler.END
    db.set_profile(height, age, gender, pal)
    await query.edit_message_text(
        f"👤 Profile saved: {height} cm, {age} yr, {gender}, PAL {pal}"
    )
    context.user_data.clear()
    # Auto-export
    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", "Auto-export: profile update"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)
    return ConversationHandler.END


async def _save_profile_msg(update: Update, context: ContextTypes.DEFAULT_TYPE, pal: float) -> int:
    d = context.user_data
    height = d.get("_profile_height")
    age = d.get("_profile_age")
    gender = d.get("_profile_gender")
    if None in (height, age, gender):
        await update.message.reply_text("Something went wrong. Start over with ⚙️ Settings.")
        context.user_data.clear()
        return ConversationHandler.END
    db.set_profile(height, age, gender, pal)
    await update.message.reply_text(
        f"👤 Profile saved: {height} cm, {age} yr, {gender}, PAL {pal}",
        reply_markup=DEFAULT_KEYBOARD,
    )
    context.user_data.clear()
    # Auto-export
    try:
        import subprocess
        subprocess.run(["python3", "export_json.py"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "add", "docs/data/workouts.json"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
        subprocess.run(["git", "commit", "-m", "Auto-export: profile update"], cwd="/Users/billkim/gym-tracker", check=False, capture_output=True)
        subprocess.run(["git", "push", "origin", "main"], cwd="/Users/billkim/gym-tracker", check=True, capture_output=True)
    except Exception as e:
        logger.error("Auto-export/push failed: %s", e)
    return ConversationHandler.END


async def rest_day_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    today = _today()
    if db.is_rest_day(today):
        await update.message.reply_text(
            "🛌 Today is already marked as a rest day.", reply_markup=DEFAULT_KEYBOARD
        )
        return
    db.mark_rest_day(today)
    streak = db.get_streak()
    msg = "🛌 Rest day marked for today."
    if streak > 1:
        msg += f"\n🔥 Streak alive: *{streak}* days"
    await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=DEFAULT_KEYBOARD)


async def today_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    today_rows = db.get_workouts_for_date(_today())
    avg7 = db.get_7day_avg()
    await update.message.reply_text(
        fmt_report(today_rows, avg7), parse_mode="Markdown", reply_markup=DEFAULT_KEYBOARD
    )


async def week_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not _authorized(update):
        return
    avg7 = db.get_7day_avg()
    recent = db.get_recent_workouts(7)
    lines = ["📅 *Last 7 Days*"]
    for w in recent:
        lines.append(
            f"  • {w['date']} {w['day']} — {w['machine']} {w['calories']} kcal"
        )
    lines.append("")
    lines.append(f"7-day avg: *{avg7['avg_cal']}* kcal/day, *{round(avg7['avg_min'])}* min/day")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def daily_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    today_rows = db.get_workouts_for_date(_today())
    if not today_rows and not db.is_rest_day(_today()):
        return
    avg7 = db.get_7day_avg()
    text = fmt_report(today_rows, avg7)
    await context.bot.send_message(
        chat_id=config.USER_ID, text=text, parse_mode="Markdown", reply_markup=DEFAULT_KEYBOARD
    )


async def beat_yesterday_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    yesterday = db.get_yesterday_summary()
    today_so_far = db.get_today_summary()
    lines = ["🌅 *Morning Challenge*"]
    lines.append(f"Yesterday: *{yesterday['total_cal']}* kcal, *{round(yesterday['total_min'])}* min")
    if today_so_far["total_cal"] > 0:
        lines.append(f"So far today: *{today_so_far['total_cal']}* kcal")
    if yesterday["total_cal"] > 0:
        need = round(yesterday["total_cal"] - today_so_far["total_cal"], 1)
        if need > 0:
            lines.append(f"🎯 Beat yesterday: need *{need}* more kcal")
        else:
            lines.append("🎯 Already beat yesterday!")
    else:
        lines.append("No workout yesterday. Today is a fresh start.")
    await context.bot.send_message(
        chat_id=config.USER_ID, text="\n".join(lines), parse_mode="Markdown", reply_markup=DEFAULT_KEYBOARD
    )


async def weekly_alert(context: ContextTypes.DEFAULT_TYPE) -> None:
    summary = db.get_week_summary()
    target_cal = config.DAILY_GOAL_KCAL * 7
    target_min = summary["target_min"]
    cal_gap = round(target_cal - summary["total_cal"], 1)
    min_gap = round(target_min - summary["total_min"], 1)
    lines = [f"📅 *Week of {summary['week_start']}*"]
    lines.append(f"  • Calories: *{summary['total_cal']}* / {target_cal} kcal")
    lines.append(f"  • Cardio time: *{round(summary['total_min'])}* / {target_min} min")
    lines.append(f"  • Workouts: *{summary['workout_count']}* on *{summary['days_with_workouts']}* days")
    lines.append("")
    if cal_gap > 0 and min_gap > 0:
        lines.append(f"⚠️ Need *{cal_gap}* kcal and *{min_gap}* min to hit weekly targets.")
    elif cal_gap > 0:
        lines.append(f"⚠️ Need *{cal_gap}* kcal to hit weekly calorie target.")
    elif min_gap > 0:
        lines.append(f"⚠️ Need *{min_gap}* min to hit weekly cardio target.")
    else:
        lines.append("🎯 Weekly targets reached!")
    await context.bot.send_message(
        chat_id=config.USER_ID, text="\n".join(lines), parse_mode="Markdown", reply_markup=DEFAULT_KEYBOARD
    )


async def daily_weight_math_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    msg = compute_weight_math()
    if msg:
        await context.bot.send_message(
            chat_id=config.USER_ID, text=msg, parse_mode="Markdown", reply_markup=DEFAULT_KEYBOARD
        )


async def daily_gmail_brief(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Run the standalone Gmail brief script. It handles its own
    noise filtering, deduplication, and Telegram sending."""
    import subprocess
    import sys
    try:
        proc = subprocess.run(
            [sys.executable, "scripts/gmail_brief.py"],
            cwd="/Users/billkim/gym-tracker",
            capture_output=True,
            text=True,
            timeout=60,
        )
        if proc.returncode != 0:
            logger.error("daily_gmail_brief failed: %s", proc.stderr)
        else:
            out = proc.stdout.strip()
            if out and not out.startswith("[SKIP]"):
                logger.info("daily_gmail_brief sent: %s", out[:200])
            else:
                logger.info("daily_gmail_brief: %s", out or "no output")
    except Exception as e:
        logger.error("daily_gmail_brief exception: %s", e)


def main() -> None:
    db.init_db()
    application = Application.builder().token(config.BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[
            CommandHandler("log", log_cmd),
            MessageHandler(filters.Regex("^(Log Workout)$"), log_cmd),
            MessageHandler(filters.PHOTO, photo_received),
        ],
        states={
            TYPE: [CallbackQueryHandler(type_selected)],
            DURATION: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, duration_received),
                CallbackQueryHandler(duration_skip, pattern="^skip$"),
            ],
            CALORIES: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, calories_received),
                CallbackQueryHandler(calories_skip, pattern="^skip$"),
            ],
            DISTANCE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, distance_received),
                CallbackQueryHandler(distance_skip, pattern="^skip$"),
            ],
            CONFIRM: [CallbackQueryHandler(confirm_handler)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    weight_conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^(⚖️ Weight)$"), weight_button_start),
        ],
        states={
            WAITING_WEIGHT: [MessageHandler(filters.TEXT, weight_received)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("today", today_cmd))
    application.add_handler(CommandHandler("week", week_cmd))
    application.add_handler(CommandHandler("rest", rest_day_cmd))
    application.add_handler(CommandHandler("export", export_cmd))
    application.add_handler(CommandHandler("goal", goal_cmd))
    application.add_handler(CommandHandler("weight", weight_cmd))
    application.add_handler(CommandHandler("me", me_cmd))
    application.add_handler(CommandHandler("target", target_cmd))
    application.add_handler(CommandHandler("math", math_cmd))
    application.add_handler(MessageHandler(filters.Regex("^(🛌 Rest Day)$"), rest_day_cmd))
    application.add_handler(MessageHandler(filters.Regex("^(⚙️ Settings)$"), settings_cmd))

    # Settings conversations
    settings_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(target_weight_start, pattern="^set_target$")],
        states={
            TARGET_WEIGHT: [
                CallbackQueryHandler(target_weight_received, pattern="^tw_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, target_weight_text),
            ],
            TARGET_DATE: [
                CallbackQueryHandler(target_date_received, pattern="^td_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, target_date_text),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    profile_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(profile_start, pattern="^set_profile$")],
        states={
            PROFILE_HEIGHT: [
                CallbackQueryHandler(profile_height_received, pattern="^ph_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, profile_height_text),
            ],
            PROFILE_AGE: [
                CallbackQueryHandler(profile_age_received, pattern="^pa_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, profile_age_text),
            ],
            PROFILE_GENDER: [
                CallbackQueryHandler(profile_gender_received, pattern="^pg_"),
            ],
            PROFILE_PAL: [
                CallbackQueryHandler(profile_pal_received, pattern="^pp_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, profile_pal_text),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    # Main menu callback (math button only — set_target/set_profile are conversation entry points)
    application.add_handler(CallbackQueryHandler(settings_callback, pattern="^math$"))
    application.add_handler(settings_conv)
    application.add_handler(profile_conv)
    application.add_handler(weight_conv)
    application.add_handler(conv)

    job_queue = application.job_queue
    job_queue.run_daily(daily_report, time=datetime.time(hour=config.DAILY_REPORT_HOUR, minute=0))
    # Morning challenge — Mon/Wed/Fri 8:10 AM CDT
    job_queue.run_daily(beat_yesterday_report, time=datetime.time(hour=8, minute=10), days=(0, 2, 4))
    # Daily Gmail brief — 9:00 AM CDT
    job_queue.run_daily(daily_gmail_brief, time=datetime.time(hour=9, minute=0))
    # Saturday 9:00 AM CDT (Mini local time)
    job_queue.run_daily(weekly_alert, time=datetime.time(hour=9, minute=0), days=(5,))
    # Morning weight math — 8:00 AM CDT
    job_queue.run_daily(daily_weight_math_report, time=datetime.time(hour=8, minute=0))

    application.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
